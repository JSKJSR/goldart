# routes/export.py — swap exporters independently
from flask import Blueprint, send_file
from db.queries import get_all_trades, get_stats_summary
import io
from datetime import date

export_bp = Blueprint("export", __name__)


@export_bp.get("/excel")
def to_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    trades = get_all_trades(limit=500)
    wb     = openpyxl.Workbook()
    ws     = wb.active
    ws.title = "Trades"

    headers = [
        "ID", "Date", "Time", "Direction", "4H Bias", "1H Bias",
        "Entry", "SL", "TP", "Lot Size", "Exit", "Result",
        "PnL", "RR", "Checklist", "Rating", "Emotion", "Notes",
    ]

    gold  = "FFD700"
    black = "1A1A1A"

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = Font(bold=True, color=black)
        cell.fill      = PatternFill("solid", fgColor=gold)
        cell.alignment = Alignment(horizontal="center")

    for row, t in enumerate(trades, 2):
        ws.append([
            t["id"], t["date"], t["time"], t["direction"],
            t["bias_4h"], t["bias_1h"], t["entry_price"], t["sl_price"],
            t["tp_price"], t["lot_size"], t["exit_price"], t["result"],
            t["pnl"], t["rr_achieved"], t["checklist_score"],
            t["setup_rating"], t["emotion"], t["notes"],
        ])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"goldart_trades_{date.today().isoformat()}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@export_bp.get("/pdf")
def to_pdf():
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    trades = get_all_trades(limit=200)
    stats  = get_stats_summary()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), title="GOLDART Report")
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("GOLDART — Trade Report", styles["Title"]))
    elements.append(Spacer(1, 12))

    summary = [
        ["Total Trades", "Win Rate", "Avg RR", "Total PnL"],
        [
            stats.get("total", 0),
            f"{stats.get('win_rate', 0)}%",
            stats.get("avg_rr", 0),
            f"${stats.get('total_pnl', 0)}",
        ],
    ]
    st = Table(summary)
    st.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFD700")),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN",      (0, 0), (-1, -1), "CENTER"),
        ("GRID",       (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(st)
    elements.append(Spacer(1, 20))

    rows = [["Date", "Dir", "Entry", "SL", "TP", "Lot", "Result", "PnL", "RR"]]
    for t in trades:
        rows.append([
            t["date"], t["direction"], t["entry_price"], t["sl_price"],
            t["tp_price"], t["lot_size"], t["result"] or "OPEN",
            f"${t['pnl'] or 0}", t["rr_achieved"] or "-",
        ])

    tbl = Table(rows, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A1A1A")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.HexColor("#FFD700")),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",   (0, 0), (-1, -1), 8),
        ("GRID",       (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
    ]))
    elements.append(tbl)

    doc.build(elements)
    buf.seek(0)

    filename = f"goldart_report_{date.today().isoformat()}.pdf"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/pdf")
